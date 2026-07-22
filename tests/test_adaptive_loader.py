# -*- coding: utf-8 -*-
"""adaptive_loader 회귀 테스트.

Codex 지적 3건 재현·검증:
  1) 단일표형 시트가 여러 개일 때 가장 큰 시트 하나만 채택 → 나머지 누락
  2) Excel serial number 날짜를 pd.to_datetime이 1970 기준 나노초로 오해석
  3) 작성자/전표유형 결측을 astype(str)로 변환 → "nan" 문자열이 되어
     메타필드 존재 판정과 fraud score MAXPOSS(분모)를 왜곡
"""
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

SRC = Path(__file__).resolve().parent.parent / "src"
sys.path.insert(0, str(SRC))
from adaptive_loader import load_ledger  # noqa: E402

BASE_HEADER = ["계정코드", "계정명", "전표일자", "적요", "거래처명", "차변", "대변", "잔액"]

# fraud_screen.py의 배점(fraud_weights.json 기본값)
W_BASE = {"상대계정조합": 30, "결산수정키워드": 25, "거래처없는대형": 20, "중복전표": 15,
          "라운드금액": 15, "기말": 10, "주말": 10, "적요모호": 10}
W_META = {"수기전표": 20, "소급입력": 20, "작성자이상": 10}

EXCEL_EPOCH_2025_01_01 = 45658   # Excel serial(1899-12-30 기준) = 2025-01-01


def _add_sheet(wb, title, header, rows, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(["회사명 : 테스트회사"])
    ws.append(header)
    for r in rows:
        ws.append(r)
    return ws


def _rows(codes, date_fn, extra_fn=None):
    """계정코드 목록 → 원장 행. date_fn(i)가 전표일자 값을 만든다."""
    out = []
    for i, c in enumerate(codes):
        row = [str(c), f"계정{c}", date_fn(i), "거래내역", "거래처A", 1000 * (i + 1), 0, 0]
        if extra_fn is not None:
            row += extra_fn(i)
        out.append(row)
    return out


def _dt(i):
    return datetime(2025, 1, 1 + i)


def _save(wb, tmp_path, name="원장_2025.xlsx"):
    p = tmp_path / name
    wb.save(p)
    return p


# ───────────────────────── 지적 1: 다중 단일표형 시트 ─────────────────────────

def test_multiple_flat_sheets_same_structure_all_rows_loaded(tmp_path):
    """동일 구조의 원장 시트가 여러 개면 전부 적재돼야 한다(현재는 가장 큰 것만)."""
    # 두 시트 모두 단일표형 판정 기준(고유계정 > 5)을 넘기되 크기는 다르게 둔다.
    # 기존 구현은 max()로 큰 시트 하나만 채택하므로 작은 시트가 통째로 누락된다.
    wb = Workbook()
    _add_sheet(wb, "원장1", BASE_HEADER,
               _rows([101, 102, 103, 104, 105, 106, 107], _dt), first=True)
    _add_sheet(wb, "원장2", BASE_HEADER,
               _rows([201, 202, 203, 204, 205, 206], _dt))
    r = load_ledger(_save(wb, tmp_path))

    assert r["meta"]["n_rows"] == 13, "두 시트의 행이 모두 적재돼야 함"
    assert r["meta"]["n_accounts"] == 13
    codes = set(r["gl"]["계정코드"])
    assert {"101", "201"} <= codes, "두 번째 시트 계정이 누락되면 안 됨"


def test_flat_sheets_with_different_structure_not_merged(tmp_path):
    """구조가 다른 시트(요약 등)는 병합되지 않아야 한다(중복 집계 방지)."""
    wb = Workbook()
    _add_sheet(wb, "원장1", BASE_HEADER, _rows([101, 102, 103, 104, 105, 106], _dt), first=True)
    _add_sheet(wb, "원장2", BASE_HEADER, _rows([201, 202, 203, 204, 205, 206], _dt))
    summary_header = ["계정코드", "계정명", "차변", "대변", "잔액", "전표일자", "비고"]
    summary_rows = [[str(c), f"계정{c}", 999, 0, 0, _dt(i), "요약"]
                    for i, c in enumerate([301, 302, 303, 304, 305, 306])]
    _add_sheet(wb, "요약", summary_header, summary_rows)

    r = load_ledger(_save(wb, tmp_path))
    codes = set(r["gl"]["계정코드"])
    assert {"101", "201"} <= codes, "동일 구조 원장 2개는 병합돼야 함"
    assert not any(c.startswith("30") for c in codes), "구조가 다른 요약 시트는 병합 금지"
    assert r["meta"]["n_rows"] == 12


# ───────────────────────── 지적 2: 날짜 형식 ─────────────────────────

def test_dates_as_excel_serial_number(tmp_path):
    """Excel serial number 날짜가 1970이 아니라 실제 연도로 파싱돼야 한다."""
    wb = Workbook()
    _add_sheet(wb, "원장", BASE_HEADER,
               _rows([101, 102, 103, 104, 105, 106],
                     lambda i: EXCEL_EPOCH_2025_01_01 + i), first=True)
    r = load_ledger(_save(wb, tmp_path))

    years = set(pd.to_datetime(r["gl"]["전표일자"]).dt.year)
    assert years == {2025}, f"serial 날짜 오해석(1970 등): {years}"
    assert r["gl"]["전표일자"].min() == pd.Timestamp("2025-01-01")


def test_dates_as_datetime(tmp_path):
    """기존 동작 회귀: datetime 셀."""
    wb = Workbook()
    _add_sheet(wb, "원장", BASE_HEADER,
               _rows([101, 102, 103, 104, 105, 106], _dt), first=True)
    r = load_ledger(_save(wb, tmp_path))
    assert set(pd.to_datetime(r["gl"]["전표일자"]).dt.year) == {2025}
    assert r["meta"]["n_rows"] == 6


def test_dates_as_string(tmp_path):
    """기존 동작 회귀: 문자열 날짜."""
    wb = Workbook()
    _add_sheet(wb, "원장", BASE_HEADER,
               _rows([101, 102, 103, 104, 105, 106],
                     lambda i: f"2025-01-{i + 1:02d}"), first=True)
    r = load_ledger(_save(wb, tmp_path))
    assert set(pd.to_datetime(r["gl"]["전표일자"]).dt.year) == {2025}
    assert r["meta"]["n_rows"] == 6


def test_dates_as_month_day_string_uses_filename_year(tmp_path):
    """기존 동작 회귀: 연도 없는 MM-DD는 파일명 연도를 붙인다."""
    wb = Workbook()
    _add_sheet(wb, "원장", BASE_HEADER,
               _rows([101, 102, 103, 104, 105, 106],
                     lambda i: f"01-{i + 1:02d}"), first=True)
    r = load_ledger(_save(wb, tmp_path, "원장_2025.xlsx"))
    assert set(pd.to_datetime(r["gl"]["전표일자"]).dt.year) == {2025}


# ───────────────────────── 지적 3: 메타필드 결측 ─────────────────────────

META_HEADER = BASE_HEADER + ["작성자", "전표유형"]


def _meta_gate(gl, col):
    """fraud_screen.py의 활성화 게이트와 동일한 판정."""
    return col in gl.columns and gl[col].notna().any()


def _maxposs(gl):
    """fraud_screen.py의 MAXPOSS(분모) 계산 재현."""
    w = dict(W_BASE)
    if _meta_gate(gl, "전표유형"):
        w["수기전표"] = W_META["수기전표"]
    if "입력일" in gl.columns and pd.to_datetime(gl["입력일"], errors="coerce").notna().any():
        w["소급입력"] = W_META["소급입력"]
    if _meta_gate(gl, "작성자"):
        w["작성자이상"] = W_META["작성자이상"]
    return sum(w.values())


def test_author_jtype_all_missing_not_treated_as_present(tmp_path):
    """작성자·전표유형 열이 있으나 전부 결측이면 '존재'로 보면 안 된다."""
    wb = Workbook()
    _add_sheet(wb, "원장", META_HEADER,
               _rows([101, 102, 103, 104, 105, 106], _dt,
                     extra_fn=lambda i: [None, None]), first=True)
    r = load_ledger(_save(wb, tmp_path))

    assert "작성자" not in r["meta"]["meta_fields"], "전결측인데 메타필드로 보고됨"
    assert "전표유형" not in r["meta"]["meta_fields"]
    assert "작성자" not in r["gl"].columns, "전결측 열은 출력에서 제외돼야 함"


def test_maxposs_denominator_not_inflated_when_meta_all_missing(tmp_path):
    """전결측 메타 때문에 MAXPOSS(분모)가 부풀지 않아야 한다."""
    wb = Workbook()
    _add_sheet(wb, "원장", META_HEADER,
               _rows([101, 102, 103, 104, 105, 106], _dt,
                     extra_fn=lambda i: [None, None]), first=True)
    gl = load_ledger(_save(wb, tmp_path))["gl"]

    assert _meta_gate(gl, "작성자") is False
    assert _meta_gate(gl, "전표유형") is False
    assert _maxposs(gl) == sum(W_BASE.values()) == 135, "분모가 기본 135여야 함"


# ───────────────── 지적(2차) 3: 시트별형 합성 계정코드 ↔ TB 정합 ─────────────────

# 시트별형: 시트 1개 = 계정 1개. 계정코드 열이 없고 계정명 열만 있는 형식.
NAMEONLY_HEADER = ["계정명", "전표일자", "적요", "거래처명", "차변", "대변", "잔액"]


def _add_account_sheet(wb, title, acct_name, first=False, n=3):
    rows = [[acct_name, _dt(i), "거래내역", "거래처A", 1000 * (i + 1), 0, 0]
            for i in range(n)]
    return _add_sheet(wb, title, NAMEONLY_HEADER, rows, first=first)


def test_sheetwise_synthetic_code_applied_to_tb(tmp_path):
    """계정코드 열이 없어 GL에 합성코드를 부여하면 TB에도 같은 매핑이 적용돼야 한다.
       (fr02_validate.py가 tb.merge(det, on='계정코드')로 대사하므로 불일치 시 조인 실패)"""
    wb = Workbook()
    _add_account_sheet(wb, "현금", "현금", first=True)
    _add_account_sheet(wb, "보통예금", "보통예금")
    _add_account_sheet(wb, "매출", "매출")
    r = load_ledger(_save(wb, tmp_path))
    gl, tb = r["gl"], r["tb"]

    assert r["meta"]["layout"] == "시트별형"
    assert set(gl["계정코드"]) == set(tb["계정코드"]), "GL·TB 계정코드가 어긋남"

    # fr02의 역산 대사 조인 재현 — 매칭 실패(NaN)가 없어야 한다.
    det = gl.groupby("계정코드").agg(상세차변=("차변", "sum")).reset_index()
    m = tb.merge(det, on="계정코드", how="left")
    assert m["상세차변"].notna().all(), "TB-GL 조인 실패(상세합이 붙지 않음)"
    assert m["상세차변"].sum() == gl["차변"].sum(), "조인 후 금액이 보존되지 않음"


def test_sheetwise_with_numeric_code_still_matches(tmp_path):
    """계정코드가 있는 경우(회귀): 합성 매핑 없이도 GL-TB가 일치해야 한다."""
    header = ["계정코드", "계정명", "전표일자", "적요", "거래처명", "차변", "대변", "잔액"]
    wb = Workbook()
    for i, (code, name) in enumerate([("101", "현금"), ("102", "보통예금"), ("401", "매출")]):
        rows = [[code, name, _dt(j), "거래내역", "거래처A", 1000, 0, 0] for j in range(3)]
        _add_sheet(wb, name, header, rows, first=(i == 0))
    r = load_ledger(_save(wb, tmp_path))
    gl, tb = r["gl"], r["tb"]

    assert set(gl["계정코드"]) == set(tb["계정코드"])
    det = gl.groupby("계정코드").agg(상세차변=("차변", "sum")).reset_index()
    m = tb.merge(det, on="계정코드", how="left")
    assert m["상세차변"].notna().all()


def test_sheetwise_name_variants_map_consistently(tmp_path):
    """계정명에 공백·대소문자 차이가 있어도 GL과 TB 매핑이 서로 어긋나지 않아야 한다."""
    wb = Workbook()
    _add_account_sheet(wb, "s1", "현금", first=True)
    _add_account_sheet(wb, "s2", "현금 ")        # 뒤 공백
    _add_account_sheet(wb, "s3", "Cash")
    _add_account_sheet(wb, "s4", "cash")         # 대소문자 차이
    r = load_ledger(_save(wb, tmp_path))
    gl, tb = r["gl"], r["tb"]

    assert set(gl["계정코드"]) == set(tb["계정코드"]), "이름 변형 시 GL·TB 매핑 불일치"
    det = gl.groupby("계정코드").agg(상세차변=("차변", "sum")).reset_index()
    m = tb.merge(det, on="계정코드", how="left")
    assert m["상세차변"].notna().all(), "이름 변형 계정에서 조인 실패"


def test_author_partial_missing_keeps_na_not_nan_string(tmp_path):
    """일부 행만 작성자가 있으면, 결측은 NaN이어야 하고 'nan' 문자열이면 안 된다."""
    wb = Workbook()
    _add_sheet(wb, "원장", META_HEADER,
               _rows([101, 102, 103, 104, 105, 106], _dt,
                     extra_fn=lambda i: (["김담당", "수기"] if i < 2 else [None, None])),
               first=True)
    r = load_ledger(_save(wb, tmp_path))
    gl = r["gl"]

    assert "작성자" in r["meta"]["meta_fields"], "값이 일부라도 있으면 존재로 봐야 함"
    assert "nan" not in set(gl["작성자"].dropna().astype(str)), "'nan' 문자열이 값으로 남음"
    assert gl["작성자"].notna().sum() == 2, "실제 값이 있는 행만 non-null이어야 함"
    assert gl["작성자"].isna().sum() == 4
